import os, json, re, io, anthropic, hashlib, pickle
from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from werkzeug.utils import secure_filename
import pdfplumber, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'mca-analyzer-secret-2026')
app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', '/tmp/uploads')
app.config['HISTORY_FOLDER'] = os.environ.get('HISTORY_FOLDER', '/tmp/history')
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024
ALLOWED_EXTENSIONS = {'pdf', 'csv', 'txt'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['HISTORY_FOLDER'], exist_ok=True)

# Login credentials - set via environment variables
USERS = {
    os.environ.get('USERNAME1', 'dave'): os.environ.get('PASSWORD1', 'mca2026'),
    os.environ.get('USERNAME2', 'admin'): os.environ.get('PASSWORD2', 'analyze2026'),
}

def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def allowed_file(f): return '.' in f and f.rsplit('.',1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(path):
    text = ""
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: text += t + "\n"
    return text

def extract_text(path):
    ext = path.rsplit('.',1)[1].lower()
    if ext == 'pdf': return extract_text_from_pdf(path)
    with open(path,'r',encoding='utf-8',errors='ignore') as f: return f.read()

def calc_monthly(amount, frequency):
    freq = (frequency or 'weekly').lower()
    if 'daily' in freq: return amount * 22
    if 'bi' in freq: return amount * 2
    if 'monthly' in freq: return amount * 1
    return amount * 4

def save_history(company_name, data, excel_bytes):
    safe = re.sub(r'[^\w\s-]','',company_name).strip().replace(' ','_')
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    entry_id = "{}_{}".format(safe, ts)
    history_path = os.path.join(app.config['HISTORY_FOLDER'], entry_id + '.pkl')
    with open(history_path, 'wb') as f:
        pickle.dump({'id': entry_id, 'company_name': company_name, 'data': data,
                     'excel': excel_bytes, 'timestamp': ts}, f)
    return entry_id

def load_history():
    entries = []
    folder = app.config['HISTORY_FOLDER']
    for fname in sorted(os.listdir(folder), reverse=True)[:10]:
        if fname.endswith('.pkl'):
            try:
                with open(os.path.join(folder, fname), 'rb') as f:
                    entry = pickle.load(f)
                    entries.append({'id': entry['id'], 'company_name': entry['company_name'],
                                   'timestamp': entry['timestamp']})
            except: pass
    return entries

def load_entry(entry_id):
    path = os.path.join(app.config['HISTORY_FOLDER'], entry_id + '.pkl')
    if not os.path.exists(path): return None
    with open(path, 'rb') as f:
        return pickle.load(f)

def parse_with_claude(raw_text, company_name=""):
    client = anthropic.Anthropic()
    cn = company_name if company_name else 'auto-detect'
    prompt = (
        "You are an expert MCA underwriter analyzing bank statements. You are the LENDER.\n"
        "Return ONLY valid JSON, no markdown, no explanation.\n\n"
        "BANK STATEMENT TEXT (may contain multiple months):\n"
        + raw_text[:80000] +
        "\n\nCRITICAL INSTRUCTIONS:\n"
        "1. Find EVERY statement period. Look for CHECKING SUMMARY headers and date ranges.\n"
        "2. Extract ALL months found - do not skip any.\n"
        "3. Most recent partial month gets is_mtd = true.\n\n"
        "FOR EACH MONTH:\n"
        "- total_deposits: Deposits and Additions total from CHECKING SUMMARY\n"
        "- true_deposits: total_deposits minus MCA funding Fedwires (B/O field shows lender names) and minus Online Transfer From Chk entries\n"
        "- Shileno LLC wires are real client payments - KEEP in true deposits\n"
        "- adb: average of all values in DAILY ENDING BALANCE table\n"
        "- neg_days: count of negative balances in DAILY ENDING BALANCE table\n"
        "- days_below_1000: count of balances under 1000 in DAILY ENDING BALANCE table\n"
        "- funding_events: incoming Fedwire credits that are MCA loans\n\n"
        "FOR CURRENT POSITIONS: find all recurring ACH debits in Electronic Withdrawals.\n"
        "- amount = the recurring payment amount\n"
        "- frequency = daily, weekly, bi-weekly, or monthly\n"
        "- Use the most recent payment amount per lender\n\n"
        "Return this JSON:\n"
        '{"company_name":"string","account_number_last4":"string","num_bank_accounts":1,'
        '"offer_decline":"DECLINE","holdback_pct":0.0,"sos_info":"","court_search_notes":"",'
        '"account_notes":[],'
        '"current_positions":[{"lender":"name","amount":0.0,"frequency":"weekly","notes":""}],'
        '"months":[{"month_label":"Mon-YY","period":"MM/DD to MM/DD","is_mtd":false,'
        '"total_deposits":0.0,"true_deposits":0.0,"true_deposit_notes":"",'
        '"neg_days":0,"nsf_count":0,"od_count":0,"num_transactions":0,'
        '"adb":0.0,"days_below_1000":0,'
        '"funding_events":[{"funder":"name","amount":0.0,"date":"MM/DD"}],'
        '"notes":""}]}\n\n'
        "Company name if provided: " + cn
    )
    msg = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=8000,
        messages=[{"role":"user","content":prompt}]
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r'^```json\s*','',raw)
    raw = re.sub(r'^```\s*','',raw)
    raw = re.sub(r'\s*```$','',raw)
    data = json.loads(raw)
    total = 0
    for pos in data.get("current_positions", []):
        monthly = calc_monthly(pos.get("amount", 0), pos.get("frequency", "weekly"))
        pos["monthly_amount"] = monthly
        total += monthly
    data["total_current_positions"] = total
    return data

def merge_data(existing, new_data):
    existing_labels = {m['month_label'] for m in existing.get('months', [])}
    for m in new_data.get('months', []):
        if m['month_label'] not in existing_labels:
            existing['months'].append(m)
    existing['months'].sort(key=lambda x: x.get('month_label',''), reverse=True)
    existing_lenders = {p['lender'] for p in existing.get('current_positions', [])}
    for p in new_data.get('current_positions', []):
        if p['lender'] not in existing_lenders:
            existing['current_positions'].append(p)
    total = sum(p.get('monthly_amount', calc_monthly(p.get('amount',0), p.get('frequency','weekly')))
                for p in existing.get('current_positions', []))
    existing['total_current_positions'] = total
    return existing

def build_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analysis"

    GOLD_PALE="1A1A2E"; LIGHT_YELLOW="16213E"; DARK_GOLD="E2B96F"
    GREEN_BG="1B4332"; GREEN_FG="95D5B2"; RED_BG="4A1520"; RED_FG="FF6B6B"
    BLUE="7EB8F7"; PURPLE_BG="2D1B69"; GRAY="1E1E2E"
    NEG_BG="4A1520"; OK_BG="1B4332"; MONTH_BG="2A1F00"
    MONTH_FG="E2B96F"; WHITE="E8E8F0"

    thin = Side(style='thin')
    def border_all():
        return Border(left=thin,right=thin,top=thin,bottom=thin)

    def w(row,col,value="",bold=False,sz=10,color=None,bg=None,align="left",
          bdr=False,italic=False,ul=False,wrap=False):
        c = ws.cell(row=row,column=col,value=value)
        kw={"bold":bold,"size":sz,"italic":italic}
        if ul: kw["underline"]="single"
        kw["color"] = color if color else WHITE
        c.font=Font(**kw)
        if bg: c.fill=PatternFill("solid",start_color=bg)
        c.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
        if bdr: c.border=border_all()
        return c

    def merge(r1,c1,r2,c2): ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)

    for col,wd in {1:3,2:34,3:20,4:16,5:16,6:14,7:14,8:36}.items():
        ws.column_dimensions[get_column_letter(col)].width=wd

    ws.sheet_view.showGridLines = False

    row=1
    ws.row_dimensions[row].height=6; row+=1

    ws.row_dimensions[row].height=26
    w(row,3,"Amounts ($) / No.",bold=True,sz=9,align="center",bg=GRAY)
    w(row,4,"frequency",sz=9,align="center",bg=GRAY,italic=True)
    merge(row,5,row,5)
    c=ws.cell(row=row,column=5,value="APPROVED")
    c.font=Font(bold=True,size=10,color=GREEN_FG)
    c.fill=PatternFill("solid",start_color=GREEN_BG)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=border_all()
    w(row,6,"☑" if data.get("offer_decline")=="OFFER" else "☐",sz=14,align="center",bg=GREEN_BG,color=GREEN_FG)
    merge(row,7,row+1,8)
    c=ws.cell(row=row,column=7,value="Update Sheet\nTab Color")
    c.font=Font(bold=True,size=11,color=DARK_GOLD)
    c.fill=PatternFill("solid",start_color=PURPLE_BG)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    row+=1

    ws.row_dimensions[row].height=22
    c=ws.cell(row=row,column=5,value="DECLINED")
    c.font=Font(bold=True,size=10,color=RED_FG)
    c.fill=PatternFill("solid",start_color=RED_BG)
    c.alignment=Alignment(horizontal="center",vertical="center")
    c.border=border_all()
    w(row,6,"☑" if data.get("offer_decline")=="DECLINE" else "☐",sz=14,align="center",bg=RED_BG,color=RED_FG)
    row+=1

    ws.row_dimensions[row].height=6; row+=1
    ws.row_dimensions[row].height=18
    w(row,5,"☐",sz=16,align="center"); row+=1

    ws.row_dimensions[row].height=24
    merge(row,2,row,6)
    c=ws.cell(row=row,column=2,value=data.get("company_name","COMPANY NAME").upper())
    c.font=Font(bold=True,size=13,color=DARK_GOLD)
    c.fill=PatternFill("solid",start_color="0D0D1A")
    c.alignment=Alignment(horizontal="left",vertical="center")
    row+=1

    ws.row_dimensions[row].height=20
    w(row,2,"OFFER / DECLINE",bold=True,ul=True,sz=10,color=DARK_GOLD)
    w(row,3,"$0.00",sz=10,color=BLUE)
    w(row,4,"daily",sz=9,italic=True)
    w(row,5,"No. of Bank Accts",bold=True,sz=9)
    w(row,6,data.get("num_bank_accounts",1),bold=True,sz=11,align="center",color=DARK_GOLD)
    row+=1

    for note in data.get("account_notes",[])[:4]:
        ws.row_dimensions[row].height=15
        clr=RED_FG if any(x in note.lower() for x in ["1,000","negative","nsf"]) else WHITE
        w(row,2,"*"+note,sz=9,italic=True,color=clr); row+=1

    for _ in range(max(0,3-len(data.get("account_notes",[])))):
        ws.row_dimensions[row].height=14; row+=1

    ws.row_dimensions[row].height=18
    w(row,3,"Holdback %",bold=True,sz=10,align="right",color=WHITE)
    hb=data.get("holdback_pct",0)
    w(row,4,"{:.2f}%".format(hb),sz=10,align="center",color=DARK_GOLD); row+=1

    ws.row_dimensions[row].height=18
    w(row,2,"SOS",bold=True,ul=True,sz=10,color=DARK_GOLD)
    w(row,3,"New Holdback %",bold=True,sz=10,align="right",color=WHITE)
    w(row,4,"{:.2f}%".format(hb),sz=10,align="center",color=DARK_GOLD); row+=1

    ws.row_dimensions[row].height=16
    sos=data.get("sos_info","")
    w(row,2,sos if sos else "Active MM/DD/YYYY",sz=9,italic=True); row+=2

    ws.row_dimensions[row].height=18
    w(row,2,"Court Search",bold=True,ul=True,sz=10,color=DARK_GOLD); row+=1
    ws.row_dimensions[row].height=16
    court=data.get("court_search_notes","")
    w(row,2,court if court else "*No court records found",sz=9,italic=True,wrap=True); row+=2

    ws.row_dimensions[row].height=20
    acct=data.get("account_number_last4","")
    merge(row,2,row,4)
    c=ws.cell(row=row,column=2,value=acct if acct else "ACCOUNT DETAILS")
    c.font=Font(bold=True,size=12,color=DARK_GOLD)
    c.fill=PatternFill("solid",start_color="0D0D1A")
    c.alignment=Alignment(horizontal="left",vertical="center")
    row+=2

    ws.row_dimensions[row].height=18
    w(row,2,"Current Positions:",bold=True,ul=True,sz=10,color=DARK_GOLD)
    total=data.get("total_current_positions",0)
    w(row,3,"${:,.2f}".format(total) if total else "$0.00",bold=True,sz=10,color=BLUE,ul=True); row+=1

    for pos in data.get("current_positions",[]):
        ws.row_dimensions[row].height=15
        lender=pos.get("lender",""); amt=pos.get("amount",0)
        freq=pos.get("frequency","weekly"); notes=pos.get("notes","")
        w(row,2,lender,sz=9,color=BLUE)
        w(row,3,"${:,.2f}".format(amt) if amt else "",sz=9,align="right",color=WHITE)
        if freq: w(row,4,"*"+freq,sz=9,italic=True,color=WHITE)
        if notes: w(row,5,"*"+notes,sz=9,italic=True,color=RED_FG)
        row+=1

    ws.row_dimensions[row].height=16
    w(row,2,"Other Loans / Positions:",bold=True,sz=10,color=DARK_GOLD); row+=2

    for m in data.get("months",[]):
        label=m.get("month_label",""); period=m.get("period",""); is_mtd=m.get("is_mtd",False)

        ws.row_dimensions[row].height=22
        merge(row,2,row,7)
        hdr="{} (MTD) From {}".format(label,period) if is_mtd and period else label
        c=ws.cell(row=row,column=2,value=hdr)
        c.font=Font(bold=True,size=11,color=MONTH_FG)
        c.fill=PatternFill("solid",start_color=MONTH_BG)
        c.alignment=Alignment(horizontal="left",vertical="center")
        row+=1

        ws.row_dimensions[row].height=16
        td=m.get("total_deposits",0)
        w(row,2,"Total deposits:",sz=10)
        w(row,3,"${:,.2f}".format(td),sz=10,color=BLUE)
        if is_mtd: w(row,4,"*calculated *",sz=9,italic=True,color="888899")
        row+=1

        ws.row_dimensions[row].height=16
        trd=m.get("true_deposits",0)
        lbl="True deposits (MTD):" if is_mtd else "True deposits:"
        w(row,2,lbl,sz=10)
        w(row,3,"${:,.2f}".format(trd),sz=10,color=BLUE)
        ntx=m.get("num_transactions",0)
        if is_mtd and ntx: w(row,4,str(ntx),sz=10,align="center",color=BLUE)
        tnote=m.get("true_deposit_notes","")
        if tnote: w(row,5,"*incl. "+tnote,sz=8,italic=True,color="888899",wrap=True)
        row+=1

        neg=m.get("neg_days",0); nsf=m.get("nsf_count",0); od=m.get("od_count",0)
        bar_label="Neg days # {} / NSF # {} / OD # {}".format(neg,nsf,od)
        bar_bg=NEG_BG if (neg>0 or nsf>0 or od>0) else OK_BG
        bar_fg=RED_FG if (neg>0 or nsf>0 or od>0) else GREEN_FG
        merge(row,2,row,5)
        c=ws.cell(row=row,column=2,value=bar_label)
        c.font=Font(bold=True,size=9,color=bar_fg)
        c.fill=PatternFill("solid",start_color=bar_bg)
        c.alignment=Alignment(horizontal="left",vertical="center")
        row+=1

        adb=m.get("adb",0)
        w(row,2,"ADB (average daily balance)",sz=10)
        w(row,3,"${:,.2f}".format(adb),sz=10,color=BLUE)
        w(row,4,"*calculated",sz=9,italic=True,color="888899"); row+=1

        dl=m.get("days_below_1000",0)
        w(row,2,"Days below $1,000:",sz=10)
        w(row,3,str(dl),sz=10,align="center"); row+=1

        for fe in m.get("funding_events",[]):
            w(row,2,"*Funded by "+fe.get("funder",""),sz=9,italic=True,color=BLUE)
            amt_fe=fe.get("amount",0)
            w(row,3,"with an amount of ${:,.2f}".format(amt_fe) if amt_fe else "",sz=9,italic=True,color=WHITE)
            dt=fe.get("date","")
            if dt: w(row,4,"on "+dt,sz=9,italic=True,color=WHITE)
            row+=1

        mnotes=m.get("notes","")
        if mnotes:
            w(row,2,"*"+mnotes,sz=9,italic=True,color="888899",wrap=True); row+=1

        row+=2

    ws.freeze_panes="B7"
    wb.active.sheet_properties.tabColor="C8962A"
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return out

@app.route('/login', methods=['GET','POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username','').strip()
        password = request.form.get('password','').strip()
        if USERS.get(username) == password:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        error = 'Invalid username or password'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@login_required
def index():
    history = load_history()
    return render_template('index.html', history=history)

@app.route('/analyze', methods=['POST'])
@login_required
def analyze():
    if 'files' not in request.files: return jsonify({"error":"No files uploaded"}),400
    files=request.files.getlist('files')
    company_name=request.form.get('company_name','')
    entry_id=request.form.get('entry_id','')

    if not files or all(f.filename=='' for f in files): return jsonify({"error":"No files selected"}),400

    combined_text=""
    for file in files:
        if file and allowed_file(file.filename):
            fname=secure_filename(file.filename)
            fpath=os.path.join(app.config['UPLOAD_FOLDER'],fname)
            file.save(fpath)
            try:
                combined_text+="\n\n=== FILE: {} ===\n".format(fname)+extract_text(fpath)
            except Exception as e:
                return jsonify({"error":"Failed to read {}: {}".format(fname,str(e))}),500
            finally:
                if os.path.exists(fpath): os.remove(fpath)
        else: return jsonify({"error":"Unsupported file: {}".format(file.filename)}),400

    if not combined_text.strip(): return jsonify({"error":"No text extracted"}),400

    try: new_data=parse_with_claude(combined_text,company_name)
    except Exception as e: return jsonify({"error":"AI parsing failed: {}".format(str(e))}),500

    # If adding to existing entry, merge the data
    if entry_id:
        existing = load_entry(entry_id)
        if existing:
            new_data = merge_data(existing['data'], new_data)

    try:
        excel=build_excel(new_data)
        excel_bytes = excel.read()
    except Exception as e:
        return jsonify({"error":"Excel generation failed: {}".format(str(e))}),500

    # Save to history
    cn = new_data.get("company_name","Unknown")
    save_history(cn, new_data, excel_bytes)

    safe=re.sub(r'[^\w\s-]','',cn).strip().replace(' ','_')
    return send_file(io.BytesIO(excel_bytes),as_attachment=True,
                     download_name=safe+"_analysis.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/history/<entry_id>/download')
@login_required
def download_history(entry_id):
    entry = load_entry(entry_id)
    if not entry: return "Not found", 404
    safe=re.sub(r'[^\w\s-]','',entry['company_name']).strip().replace(' ','_')
    return send_file(io.BytesIO(entry['excel']),as_attachment=True,
                     download_name=safe+"_analysis.xlsx",
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/history/<entry_id>/delete', methods=['POST'])
@login_required
def delete_history(entry_id):
    path = os.path.join(app.config['HISTORY_FOLDER'], entry_id + '.pkl')
    if os.path.exists(path): os.remove(path)
    return redirect(url_for('index'))

if __name__=='__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5001)))
